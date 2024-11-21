import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from datetime import datetime
course_dict = {}
def process_student_data(file_path):
    """
    Processes an Excel file containing student data and organizes roll numbers by course codes.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        dict: A dictionary where keys are course codes and values are sorted lists of roll numbers.
    """
    # Load the Excel file, assuming the header is in the first row
    df = pd.read_excel(file_path, header=0)
    
    # Create an empty dictionary to store course codes and roll numbers
    

    # Loop through each row and populate the dictionary
    for _, row in df.iterrows():
        course_code = row['course_code']  # Access the course_code column
        rollno = row['rollno']  # Access the rollno column

        # Check if the course_code is already a key in the dictionary
        if course_code in course_dict:
            course_dict[course_code].append(rollno)  # Append the roll number to the existing list
        else:
            course_dict[course_code] = [rollno]  # Create a new list with the roll number

    # Sort each individual key's list of roll numbers
    for course_code in course_dict:
        course_dict[course_code].sort()  # Sort the list in ascending order

    return course_dict
def sort_rooms(file_path):
    """
    Reads an Excel file and sorts rooms into floor-based and LT rooms, applying specific sorting logic.

    Args:
        file_path (str): Path to the Excel file containing room data.

    Returns:
        pd.DataFrame: A sorted DataFrame with rooms organized by floors and LT logic.
    """
    # Load the Excel file
    df = pd.read_excel(file_path)

    # Separate rooms into floor-based and LT rooms
    floor_rooms = df[df['Room No.'].astype(str).str.match(r'^\d+')]  # Numeric room numbers
    lt_rooms = df[df['Room No.'].astype(str).str.startswith('LT')]  # LT room numbers

    # Sort floor rooms by the first digit of the room number and by exam capacity
    floor_rooms['Floor'] = floor_rooms['Room No.'].astype(str).str[0]  # Extract floor number
    floor_sorted = floor_rooms.sort_values(by=['Floor', 'Exam Capacity'], ascending=[False, False])

    # Sort LT rooms by floor (0 floor first, then 1 floor)
    lt_rooms['LT_Floor'] = lt_rooms['Room No.'].astype(str).str[2]  # Extract the floor number from LT (0 or 1)
    lt_sorted = lt_rooms.sort_values(by=['LT_Floor', 'Room No.'], ascending=[False, True])

    # Combine the sorted floor rooms and LT rooms
    sorted_rooms = pd.concat([floor_sorted, lt_sorted])

    # Drop helper columns used for sorting
    sorted_rooms = sorted_rooms.drop(columns=['Floor', 'LT_Floor'])
    print(sorted_rooms)
    return sorted_rooms
def process_room_capacity(df, buffer):
    """
    Processes room data, adjusts remaining capacity, and retrieves a list of room numbers.

    Args:
        file_path (str): Path to the Excel file containing room data.
        buffer (int): A buffer value (0-5) to subtract from the "Exam Capacity".

    Returns:
        tuple: A tuple containing:
            - pd.DataFrame: The updated DataFrame with "Remaining Capacity".
            - list: A list of room numbers.
    """
    if buffer < 0 or buffer > 5:
        raise ValueError("Buffer value must be between 0 and 5.")

    # Load the Excel file
    #df = pd.read_excel(file_path)

    # Add "Remaining Capacity" column and adjust based on buffer
    df['Remaining Capacity'] = df['Exam Capacity'] - buffer

    # Retrieve the list of room numbers
    room_numbers = df["Room No."].tolist()
    
    return df, room_numbers
def process_exam_timetable(file_path, course_dict):
    """
    Processes an Excel file containing an exam timetable, formats the timetable, and sorts courses based on student count.

    Args:
        file_path (str): Path to the Excel file containing the exam timetable.
        course_dict (dict): A dictionary where keys are course codes and values are lists of student roll numbers.

    Returns:
        dict: A formatted and sorted exam timetable.
    """
    # Load the Excel file
    df = pd.read_excel(file_path, skiprows=0)

    # Create the dictionary
    exam_timetable = {}
    for _, row in df.iterrows():
        date = row['Date']
        morning_courses = row['Morning']
        evening_courses = row['Evening']

        # Add morning and evening courses to the timetable
        exam_timetable[f"{date}_morning"] = morning_courses
        exam_timetable[f"{date}_evening"] = evening_courses

    # Format the timetable to wrap each course string in a list
    formatted_timetable = {}
    for key, courses in exam_timetable.items():
        formatted_timetable[key] = [courses]

    # Sort courses based on the number of students in `course_dict`
    for key, courses in formatted_timetable.items():
        if courses[0] != 'NO EXAM':  # Skip if there is no exam
            # Split courses, sort them using the length of student lists from course_dict, and rejoin
            sorted_courses = sorted(
                courses[0].split('; '),
                key=lambda course: len(course_dict.get(course, [])),
                reverse=True
            )
            formatted_timetable[key] = ['; '.join(sorted_courses)]

    return formatted_timetable
def get_density_type():
    """
    Prompts the user to select a density type: 'dense' or 'sparse'.

    Returns:
        str: The selected density type ('dense' or 'sparse').
    """
    # Ask the user for input
    density_input = input("Enter '1' for dense or '2' for sparse: ")

    # Validate input
    while density_input not in ['1', '2']:
        density_input = input("Invalid input. Please enter '1' for dense or '2' for sparse: ")

    # Determine and return the density type
    return 'dense' if density_input == '1' else 'sparse'
def allocate_students_to_rooms(course_dict, exam_timetable, room_data, output_file='exam_allocation.xlsx'):
    """
    Allocates students to rooms for exams and generates an Excel file.

    Args:
        course_dict (dict): A dictionary where keys are course codes and values are lists of students.
        exam_timetable (dict): A dictionary where keys are exam slots and values are lists of courses.
        room_data (pd.DataFrame): A DataFrame containing room details with "Remaining Capacity".
        output_file (str): The name of the output Excel file. Defaults to 'exam_allocation.xlsx'.

    Returns:
        pd.DataFrame: A DataFrame containing the allocation details.
    """
    def get_day_from_date(date_str):
        """
        Extracts the weekday name from a date string.

        Args:
            date_str (str): Date string in the format '%Y-%m-%d %H:%M:%S'.

        Returns:
            str: The name of the weekday.
        """
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            raise ValueError(f"Unexpected date format: {date_str}")
        return date_obj.strftime('%A')

    # Initialize a list to hold allocation data
    data = []
    dfx = room_data.copy(deep=True)

    # Start allocating students to rooms
    for exam_key, course_list in exam_timetable.items():
        if 'NO EXAM' in course_list:
            continue  # Skip if no exam for that slot

        dfx = room_data.copy(deep=True)  # Reset room capacities for each exam slot

        # Split the exam_key to get date and time
        date_part, time_part = exam_key.split('_')
        day_part = get_day_from_date(date_part)  # Get the weekday

        courses = course_list[0].split('; ')  # Get the courses for that exam session

        for course in courses:
            students = course_dict.get(course, [])  # Get the list of students for the course

            # Iterate over rooms and allocate students based on remaining capacity
            for i, room in dfx.iterrows():
                room_no = room['Room No.']
                remaining_capacity = room['Remaining Capacity']
                allocated_students = []

                while students and remaining_capacity > 0:
                    # Allocate students to the room until either students or capacity is exhausted
                    allocated_students.append(students.pop(0))
                    remaining_capacity -= 1

                # Update remaining capacity in the dfx DataFrame
                dfx.at[i, 'Remaining Capacity'] = remaining_capacity

                # If we have allocated students to this room, store the result
                if allocated_students:
                    data.append({
                        'Date': date_part,
                        'Day': day_part,
                        'Time': time_part,
                        'course_code': course,
                        'Room': room_no,
                        'Allocated_students_count': len(allocated_students),
                        'Roll_list': '; '.join(allocated_students)
                    })

                # Break out of the loop if all students for the course are allocated
                if not students:
                    break

    # Create a DataFrame from the collected data
    allocation_df = pd.DataFrame(data)

    # Save to Excel file
    allocation_df.to_excel(output_file, index=False)

    print(f"Excel file '{output_file}' created successfully.")
    return allocation_df
def allocate_students_sparse(course_dict, exam_timetable, room_data, output_file='exam_allocation.xlsx'):
    """
    Allocates students to rooms for exams in a sparse allocation style and generates an Excel file.

    Args:
        course_dict (dict): A dictionary where keys are course codes and values are lists of students.
        exam_timetable (dict): A dictionary where keys are exam slots and values are lists of courses.
        room_data (pd.DataFrame): A DataFrame containing room details with "Remaining Capacity".
        output_file (str): The name of the output Excel file. Defaults to 'exam_allocation.xlsx'.

    Returns:
        pd.DataFrame: A DataFrame containing the allocation details.
    """
    # Function to get the weekday name from the date
    def get_day_from_date(date_str):
        date_obj = datetime.strptime(date_str.split()[0], '%Y-%m-%d')  # Split to get just the date part
        return date_obj.strftime('%A')  # Returns the weekday name (e.g., Saturday)

    # Function to get floor from room number
    def get_floor(room_no):
        room_no_str = str(room_no)  # Ensure room_no is treated as a string
        if room_no_str.startswith('LT'):
            return 6  # Assign LT rooms to a higher floor number
        else:
            return int(room_no_str[0])  # First character as floor number

    # Initialize a list to hold the data for the DataFrame
    data = []

    # Add 'Floor' column to room_data
    room_data['Floor'] = room_data['Room No.'].apply(get_floor)

    # Sort rooms by Floor and Remaining Capacity
    room_data_sorted = room_data.sort_values(by=['Floor', 'Remaining Capacity'], ascending=[True, False]).reset_index(drop=True)

    # Drop 'Floor' column after sorting
    room_data_sorted = room_data_sorted.drop(columns=['Floor'])
    room_data = room_data_sorted

    # Calculate 'rem cap 1' and 'rem cap 2' columns
    room_data['rem cap 1'] = np.ceil(room_data['Remaining Capacity'] / 2).astype(int)  # Ceiling of remaining capacity divided by 2
    room_data['rem cap 2'] = room_data['Remaining Capacity'] - room_data['rem cap 1']  # Remaining capacity minus rem cap 1
    dfx = room_data.copy(deep=True)

    # Start allocating students to rooms
    for exam_key, course_list in exam_timetable.items():
        if 'NO EXAM' in course_list:
            continue  # Skip if no exam for that slot

        dfx = room_data.copy(deep=True)
        i = 0  # Pointer for rem cap 1
        j = 0  # Pointer for rem cap 2

        # Split the exam_key to get date and time
        date_part, time_part = exam_key.split('_')
        day_part = get_day_from_date(date_part)  # Get the weekday

        courses = course_list[0].split('; ')  # Get the courses for that exam session

        for course in courses:
            students = course_dict.get(course, [])  # Get the list of students for the course

            # Determine which pointer to use based on their values
            current_pointer = i if i <= j else j  # Choose i if i <= j, otherwise choose j
            current_rem_cap = 'rem cap 1' if current_pointer == i else 'rem cap 2'

            while students:
                allocated_students = []

                # Allocate students using the selected pointer
                if dfx.at[current_pointer, current_rem_cap] > 0:
                    # Allocate to the selected room
                    while students and dfx.at[current_pointer, current_rem_cap] > 0:
                        allocated_students.append(students.pop(0))
                        dfx.at[current_pointer, current_rem_cap] -= 1  # Decrease remaining capacity

                    # If we allocated students, store the result
                    if allocated_students:
                        data.append({
                            'Date': date_part,
                            'Day': day_part,
                            'Time': time_part,
                            'course_code': course,
                            'Room': dfx.at[current_pointer, 'Room No.'],
                            'Allocated_students_count': len(allocated_students),
                            'Roll_list': '; '.join(allocated_students)
                        })

                # After attempting to allocate, check if we need to increment the pointer
                if current_pointer == i and dfx.at[i, 'rem cap 1'] == 0:
                    i += 1  # Move to the next room for rem cap 1
                    current_pointer = i

                elif current_pointer == j and dfx.at[j, 'rem cap 2'] == 0:
                    j += 1  # Move to the next room for rem cap 2
                    current_pointer = j

    # Create a DataFrame from the collected data
    allocation_df = pd.DataFrame(data)

    # Save to Excel file
    allocation_df.to_excel(output_file, index=False)

    print(f"Excel file '{output_file}' created successfully.")
    return allocation_df
if __name__ == '__main__':
   file_path = '/content/ip_1.xlsx'  # Replace with the actual file path
   students_data = process_student_data(file_path)
   print(students_data)
   file_path = '/content/ip_3.xlsx'  # Replace with the actual file path
   df2 = sort_rooms(file_path)
   print(df2)
   file_path = '/content/ip_3.xlsx'  # Replace with the actual file path
   buffer = int(input("Enter a buffer value (0-5): "))
   try:
     updated_df, room_numbers = process_room_capacity(df2, buffer)
     print(updated_df)
     print(room_numbers)
   except ValueError as e:
     print(e)
   file_path = '/content/ip_2.xlsx'  # Replace with the actual file path
   exam_timetable = process_exam_timetable(file_path, course_dict)
   print(exam_timetable)
   density_type = get_density_type()
   print(f"You selected: {density_type}")
   if density_type == 'dense':
    # Call the function for dense allocation
    room_data=updated_df
    allocation_df = allocate_students_to_rooms(course_dict, exam_timetable, room_data)
   elif density_type == 'sparse':
    # Call the function for sparse allocation
     room_data=updated_df
     allocation_df = allocate_students_sparse(course_dict, exam_timetable, room_data)
   else:
     print("Invalid density type. Please set it to 'dense' or 'sparse'.")
   attendance_df = pd.read_excel("/content/exam_allocation.xlsx")

# Load the student data file (e.g., "student_data.csv") and convert it to a dictionary
   students_df = pd.read_excel("/content/ip_4.xlsx")
   student_dict = pd.Series(students_df.Name.values, index=students_df.Roll).to_dict()


# Create a single workbook to contain all sheets
   workbook = Workbook()
   workbook.remove(workbook.active)  # Remove the default sheet created by Workbook()

# Define a function to add a new sheet with attendance data
   def add_attendance_sheet(workbook, sheet_name, roll_numbers):
    # Add a new sheet to the workbook
     sheet = workbook.create_sheet(title=sheet_name)

    # Set the headers
     sheet['A1'] = 'Roll_No'
     sheet['B1'] = 'Name'
     sheet['C1'] = 'Signature'

    # Populate the sheet with roll numbers and corresponding names
     for idx, roll in enumerate(roll_numbers, start=2):  # Start from row 2
        sheet[f'A{idx}'] = roll.strip()
        sheet[f'B{idx}'] = student_dict.get(roll.strip(), "Unknown")  # Map roll to name
        sheet[f'C{idx}'] = ''  # Leave Signature blank


# Process each row in the attendance DataFrame
   for index, row in attendance_df.iterrows():
    # Extract information from the current row
     date = pd.to_datetime(row['Date']).strftime("%d_%m_%Y")
     course_code = row['course_code']
     room_no = row['Room']
     time_slot = row['Time'].lower()  # Ensure "morning" or "evening"

    # Create a unique sheet name using extracted information
     sheet_name = f"{date}_{course_code}_{room_no}_{time_slot}"
 
    # Get the roll list for the current row and split by comma if necessary
     roll_numbers = row['Roll_list'].split(';')

    # Add a sheet for this row's attendance data
     add_attendance_sheet(workbook, sheet_name, roll_numbers)


# Save the workbook with a descriptive name
   workbook.save("Attendance_Sheets.xlsx")